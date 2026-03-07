"""
utils/attendance_export.py — O'qituvchilar oylik davomatini Excel va PDF formatda eksport qilish
"""
import io
from datetime import datetime
from typing import List


# Status tarjimalari
STATUS_LABEL = {
    "present":  "Keldi",
    "absent":   "Kelmadi",
    "late":     "Kech keldi",
    "excused":  "Sababli",
}
STATUS_EMOJI = {
    "present": "✅",
    "absent":  "❌",
    "late":    "⏰",
    "excused": "📝",
}


def _build_pivot(records) -> tuple:
    """
    records: [{'full_name', 'date', 'status', 'comment'}, ...]
    Returns:
        teachers  — tartiblangan ism ro'yxati
        dates     — tartiblangan sana ro'yxati
        matrix    — { full_name: { date: (status, comment) } }
    """
    teachers  = sorted({r['full_name'] for r in records})
    dates     = sorted({r['date']      for r in records})
    matrix    = {t: {} for t in teachers}
    for r in records:
        matrix[r['full_name']][r['date']] = (r['status'] or 'present', r['comment'] or '')
    return teachers, dates, matrix


def generate_attendance_excel(records: list, month: str, school_name: str) -> io.BytesIO:
    """
    Oylik o'qituvchilar davomati — Excel pivot jadval.
    Ustunlar: O'qituvchi | Sana1 | Sana2 | ... | Jami ✅ | ❌ | ⏰ | 📝
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, Alignment, PatternFill,
                                  Border, Side, PatternFill)
    from openpyxl.utils import get_column_letter

    teachers, dates, matrix = _build_pivot(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLORS = {
        "present": "C6EFCE",   # yashil
        "absent":  "FFC7CE",   # qizil
        "late":    "FFEB9C",   # sariq
        "excused": "BDD7EE",   # ko'k
        "header":  "4472C4",
        "subhdr":  "D9E1F2",
        "summary": "F2F2F2",
    }

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def cell_style(cell, bold=False, align='center', color=None, font_color="000000"):
        cell.font      = Font(bold=bold, color=font_color)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border    = border
        if color:
            cell.fill  = fill(color)

    # ── Sarlavha ─────────────────────────────────────────────────
    total_cols = 1 + len(dates) + 4   # ism + sanalar + 4 statistika ustuni
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=total_cols)
    c = ws.cell(row=1, column=1,
                value=f"{school_name} — O'qituvchilar davomati — {month}")
    cell_style(c, bold=True, color=COLORS['header'], font_color="FFFFFF")
    ws.row_dimensions[1].height = 28

    # ── Header qator ─────────────────────────────────────────────
    c = ws.cell(row=2, column=1, value="O'qituvchi")
    cell_style(c, bold=True, color=COLORS['subhdr'])
    ws.column_dimensions['A'].width = 28

    for col_i, d in enumerate(dates, start=2):
        try:
            dt_obj   = datetime.strptime(d, "%Y-%m-%d")
            day_lbl  = dt_obj.strftime("%d.%m\n(%a)")
        except Exception:
            day_lbl  = d
        c = ws.cell(row=2, column=col_i, value=day_lbl)
        cell_style(c, bold=True, color=COLORS['subhdr'])
        ws.column_dimensions[get_column_letter(col_i)].width = 9
        ws.row_dimensions[2].height = 30

    # Statistika header
    for offset, label in enumerate(["✅", "❌", "⏰", "📝"], start=1):
        c = ws.cell(row=2, column=1 + len(dates) + offset, value=label)
        cell_style(c, bold=True, color=COLORS['subhdr'])
        ws.column_dimensions[get_column_letter(1 + len(dates) + offset)].width = 6

    # ── Ma'lumotlar ───────────────────────────────────────────────
    for row_i, teacher in enumerate(teachers, start=3):
        c = ws.cell(row=row_i, column=1, value=teacher)
        cell_style(c, align='left')

        counts = {"present": 0, "absent": 0, "late": 0, "excused": 0}

        for col_i, d in enumerate(dates, start=2):
            status, comment = matrix[teacher].get(d, ('present', ''))
            counts[status] = counts.get(status, 0) + 1
            label = STATUS_EMOJI.get(status, "✅")
            if comment:
                label += f"\n{comment[:20]}"
            c = ws.cell(row=row_i, column=col_i, value=label)
            cell_style(c, color=COLORS.get(status, "FFFFFF"))

        ws.row_dimensions[row_i].height = 18

        # Statistika
        for offset, key in enumerate(["present", "absent", "late", "excused"], start=1):
            c = ws.cell(row=row_i, column=1 + len(dates) + offset,
                        value=counts.get(key, 0))
            cell_style(c, color=COLORS.get(key, "FFFFFF"))

    # ── Izoh jadval ───────────────────────────────────────────────
    foot_row = 3 + len(teachers) + 1
    ws.cell(row=foot_row, column=1, value=f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_attendance_pdf(records: list, month: str, school_name: str) -> io.BytesIO:
    """
    Oylik o'qituvchilar davomati — PDF.
    Har o'qituvchi uchun: sana | status | izoh qatorlari.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontSize=14, textColor=colors.HexColor('#1f4788'),
        spaceAfter=12, alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        'Sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey,
        spaceAfter=8, alignment=TA_CENTER,
    )
    elements = []

    elements.append(Paragraph(
        f"{school_name} — O'qituvchilar oylik davomati", title_style))
    elements.append(Paragraph(f"Oy: {month}", sub_style))

    teachers, dates, matrix = _build_pivot(records)

    # Pivot jadvali — ustunlar: ism | sana | sana | ... | jami
    header_row = ["O'qituvchi"] + [
        datetime.strptime(d, "%Y-%m-%d").strftime("%d.%m") for d in dates
    ] + ["✅", "❌", "⏰", "📝"]

    table_data = [header_row]

    STATUS_COLOR_RL = {
        "present": colors.HexColor("#C6EFCE"),
        "absent":  colors.HexColor("#FFC7CE"),
        "late":    colors.HexColor("#FFEB9C"),
        "excused": colors.HexColor("#BDD7EE"),
    }

    cell_cmds = []   # TableStyle commands (bg color per cell)

    for row_i, teacher in enumerate(teachers, start=1):
        counts = {k: 0 for k in STATUS_LABEL}
        row    = [teacher]
        for col_i, d in enumerate(dates, start=1):
            status, comment = matrix[teacher].get(d, ('present', ''))
            counts[status] = counts.get(status, 0) + 1
            label = STATUS_EMOJI.get(status, "✅")
            if comment:
                label += f"\n{comment[:15]}"
            row.append(label)
            # Hujayra rangi
            cell_cmds.append(
                ('BACKGROUND', (col_i, row_i), (col_i, row_i),
                 STATUS_COLOR_RL.get(status, colors.white))
            )
        row += [counts['present'], counts['absent'], counts['late'], counts['excused']]
        table_data.append(row)

    # Ustun kengliklari
    name_w   = 5.5 * cm
    date_w   = 1.1 * cm
    stat_w   = 0.9 * cm
    col_widths = [name_w] + [date_w] * len(dates) + [stat_w] * 4

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 8),
        # Body
        ('FONTSIZE',   (0, 1), (-1, -1), 7),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',      (0, 1), (0, -1),  'LEFT'),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7f7')]),
    ] + cell_cmds))

    elements.append(tbl)
    elements.append(Spacer(1, 0.5*cm))

    # Izoh
    legend_data = [["Belgi", "Ma'no"]] + [
        [f"{STATUS_EMOJI[k]} {STATUS_LABEL[k]}", ""]
        for k in STATUS_LABEL
    ]
    legend = Table(legend_data, colWidths=[4*cm, 3*cm])
    legend.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID',     (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
    ]))
    elements.append(Paragraph("Izohlar:", styles['Normal']))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(legend)
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(
        f"<i>Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}</i>",
        styles['Normal']
    ))

    doc.build(elements)
    output.seek(0)
    return output