"""
utils/schedule_export.py — O'qituvchi jadvalini turli formatlarda eksport qilish

Formatlar:
  - Excel (.xlsx)
  - PDF
  - Rasm (.png)
"""
import io
from datetime import datetime
from typing import List, Dict


def generate_schedule_excel(slots: List[Dict]) -> io.BytesIO:
    """
    O'qituvchi jadvalini Excel formatida yaratish.
    
    Args:
        slots: Database dan olingan jadval ma'lumotlari
    
    Returns:
        BytesIO: Excel fayl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from config import WEEKDAY_LABELS
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan. pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Haftalik Jadval"
    
    # Sarlavha
    teacher_name = slots[0]['teacher_name'] if slots else "O'qituvchi"
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{teacher_name} — Haftalik dars jadvali"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 30
    
    # Header
    headers = ['№', 'Kun', 'Maktab', 'Sinf', 'Fan', 'Vaqt']
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Ma'lumotlarni guruhlab saralash
    grouped = {}
    for s in slots:
        grouped.setdefault(s['weekday'], []).append(s)
    
    # Ma'lumotlarni yozish
    row = 3
    num = 1
    for day in sorted(grouped):
        day_slots = grouped[day]
        day_label = WEEKDAY_LABELS.get(day, str(day))
        
        for slot in day_slots:
            ws.cell(row=row, column=1, value=num)
            ws.cell(row=row, column=2, value=day_label)
            ws.cell(row=row, column=3, value=slot['school_name'])
            ws.cell(row=row, column=4, value=slot['class_name'])
            ws.cell(row=row, column=5, value=slot['subject_name'])
            ws.cell(row=row, column=6, value=f"{slot['start_time']}-{slot['end_time']}")
            
            # Formatlash
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            num += 1
    
    # Ustunlar kengligini sozlash
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    
    # Faylni BytesIO ga saqlash
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_schedule_pdf(slots: List[Dict]) -> io.BytesIO:
    """
    O'qituvchi jadvalini PDF formatida yaratish.
    
    Args:
        slots: Database dan olingan jadval ma'lumotlari
    
    Returns:
        BytesIO: PDF fayl
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.enums import TA_CENTER
        from config import WEEKDAY_LABELS
    except ImportError:
        raise ImportError("reportlab kutubxonasi o'rnatilmagan. pip install reportlab")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=20,
        alignment=TA_CENTER,
    )
    
    # Sarlavha
    teacher_name = slots[0]['teacher_name'] if slots else "O'qituvchi"
    title = Paragraph(f"{teacher_name} — Haftalik dars jadvali", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Jadval ma'lumotlari
    data = [['№', 'Kun', 'Maktab', 'Sinf', 'Fan', 'Vaqt']]
    
    # Ma'lumotlarni guruhlab saralash
    grouped = {}
    for s in slots:
        grouped.setdefault(s['weekday'], []).append(s)
    
    num = 1
    for day in sorted(grouped):
        day_slots = grouped[day]
        day_label = WEEKDAY_LABELS.get(day, str(day))
        
        for slot in day_slots:
            data.append([
                str(num),
                day_label,
                slot['school_name'],
                slot['class_name'],
                slot['subject_name'],
                f"{slot['start_time']}-{slot['end_time']}"
            ])
            num += 1
    
    # Jadval yaratish
    table = Table(data, colWidths=[1.5*cm, 3*cm, 6*cm, 3*cm, 5*cm, 4*cm])
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    elements.append(table)
    
    # Sana
    elements.append(Spacer(1, 1*cm))
    date_text = Paragraph(
        f"<i>Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}</i>",
        styles['Normal']
    )
    elements.append(date_text)
    
    doc.build(elements)
    output.seek(0)
    return output


def generate_schedule_image(slots: List[Dict]) -> io.BytesIO:
    """
    O'qituvchi jadvalini rasm formatida yaratish.
    
    Args:
        slots: Database dan olingan jadval ma'lumotlari
    
    Returns:
        BytesIO: PNG rasm
    """
    try:
        from PIL import Image, ImageDraw, ImageFont
        from config import WEEKDAY_LABELS
    except ImportError:
        raise ImportError("Pillow kutubxonasi o'rnatilmagan. pip install Pillow")
    
    # Ma'lumotlarni tayyorlash
    grouped = {}
    for s in slots:
        grouped.setdefault(s['weekday'], []).append(s)
    
    # Rasm o'lchamlari
    width = 1200
    row_height = 40
    header_height = 80
    title_height = 60
    footer_height = 30
    
    total_rows = sum(len(slots) for slots in grouped.values())
    height = title_height + header_height + (total_rows * row_height) + footer_height + 40
    
    # Rasm yaratish
    img = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Font (default font)
    try:
        title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 24)
        header_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
        body_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
    except:
        title_font = ImageFont.load_default()
        header_font = ImageFont.load_default()
        body_font = ImageFont.load_default()
    
    # Sarlavha
    teacher_name = slots[0]['teacher_name'] if slots else "O'qituvchi"
    draw.rectangle([(0, 0), (width, title_height)], fill='#4472C4')
    title_text = f"{teacher_name} — Haftalik dars jadvali"
    title_bbox = draw.textbbox((0, 0), title_text, font=title_font)
    title_width = title_bbox[2] - title_bbox[0]
    draw.text(
        ((width - title_width) // 2, 18),
        title_text,
        fill='white',
        font=title_font
    )
    
    # Header
    y = title_height
    draw.rectangle([(0, y), (width, y + header_height)], fill='#D9E1F2')
    
    headers = [
        ('№', 50),
        ('Kun', 150),
        ('Maktab', 350),
        ('Sinf', 100),
        ('Fan', 250),
        ('Vaqt', 150)
    ]
    
    x_offset = 10
    for header, col_width in headers:
        draw.text((x_offset + 10, y + 25), header, fill='black', font=header_font)
        draw.line([(x_offset, y), (x_offset, y + header_height)], fill='gray', width=1)
        x_offset += col_width
    
    # Ma'lumotlar
    y += header_height
    num = 1
    
    for day in sorted(grouped):
        day_slots = grouped[day]
        day_label = WEEKDAY_LABELS.get(day, str(day))
        
        for slot in day_slots:
            # Alternating colors
            if num % 2 == 0:
                draw.rectangle([(0, y), (width, y + row_height)], fill='#f9f9f9')
            
            # Ma'lumotlar
            cells = [
                (str(num), 50),
                (day_label, 150),
                (slot['school_name'], 350),
                (slot['class_name'], 100),
                (slot['subject_name'], 250),
                (f"{slot['start_time']}-{slot['end_time']}", 150)
            ]
            
            x_offset = 10
            for text, col_width in cells:
                # Matnni qisqartirish
                if len(text) > 30:
                    text = text[:27] + '...'
                draw.text((x_offset + 10, y + 12), text, fill='black', font=body_font)
                draw.line([(x_offset, y), (x_offset, y + row_height)], fill='lightgray', width=1)
                x_offset += col_width
            
            draw.line([(0, y + row_height), (width, y + row_height)], fill='lightgray', width=1)
            y += row_height
            num += 1
    
    # Footer
    draw.text(
        (10, y + 10),
        f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        fill='gray',
        font=body_font
    )
    
    # BytesIO ga saqlash
    output = io.BytesIO()
    img.save(output, format='PNG')
    output.seek(0)
    return output